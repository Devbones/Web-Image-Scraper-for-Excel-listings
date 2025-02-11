import os
import time
import requests
from tkinter import Tk, filedialog, Label, Entry, Button, StringVar, IntVar, ttk, messagebox
from tkinter.font import Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO
import datetime
import threading
import sys


# Get the path where the bundled files are located (works for both development and packaged version)
if getattr(sys, 'frozen', False):
    # If the app is bundled (i.e., running as a standalone executable)
    app_path = sys._MEIPASS  # This is the temporary folder created by PyInstaller
else:
    # If running in a development environment (i.e., as a Python script)
    app_path = os.path.dirname(os.path.abspath(__file__))



# Path to cacert.pem
cacert_path = os.path.join(app_path, "C:/Users/kusmirek_ar/Downloads/winpython/WPy32-31230/scripts/myenv/Lib/site-packages/pip/_vendor/certifi/cacert.pem")

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
}

def main():
    def start_script():
        # Get user inputs
        excel_file = file_path.get()
        column = column_choice.get()
        start_row = int(start_row_var.get())
        end_row = int(end_row_var.get())
        output_dir = output_dir_var.get()
        target_websites = [
            target_website1_var.get(),
            target_website2_var.get(),
            target_website3_var.get(),
            target_website4_var.get()
        ]
    
        # Setup ChromeDriver
        chrome_service = Service('chromedriver.exe')
        driver = webdriver.Chrome(service=chrome_service)
    
        # Load Google
        driver.get("https://www.google.com/")
        time.sleep(2)
        # Try to find and click the "Accept all" button or "Zaakceptuj wszystko"
        try:
            # Wait for the page to load and check for the button
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))  # Wait for the body to load
    
            # Search for the "Accept all" button in English or its Polish equivalent
            accept_button = None
            try:
                # Try finding the English version "Accept all" button by button text
                accept_button = driver.find_element(By.XPATH, "//button[@id='L2AGLb']//div[contains(text(),'Accept all')]")
            except Exception:
                pass  # Continue to the next try if it's not found
    
            if not accept_button:
                try:
                    # Try finding the Polish "Zaakceptuj wszystko" button by button ID or class
                    accept_button = driver.find_element(By.XPATH, "//button[@id='L2AGLb']//div[contains(text(),'Zaakceptuj wszystko')]")
                except Exception:
                    pass  # Continue if neither button is found
    
            # If the button is found, click it
            if accept_button:
                accept_button.click()
                print("Accepted cookies pop-up.")
            else:
                print("No 'Accept all' button found.")
        except Exception as e:
            print(f"Error while handling pop-ups: {e}")
    
        # Proceed with the rest of the script
        wb = load_workbook(excel_file)
        ws = wb.active
        total_rows = end_row - start_row + 1
        processed_rows = 0
        start_time = time.time()
    
        def process_row(row_num, product_code):
            nonlocal processed_rows
            if not product_code:
                return

            # Google search and navigate to the target website
            driver.get("https://www.google.com/")
            try:
                search_box = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.NAME, "q")))

                search_box.send_keys(product_code)
                search_box.send_keys(Keys.RETURN)
            except Exception:
                return

            target_url = None
            for idx, target_website in enumerate(target_websites):
                if target_website:  # Skip empty target websites
                    try:
                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "search")))
                        links = driver.find_elements(By.XPATH, f'//a[contains(@href, "{target_website}")]')
                        if links:
                            target_url = links[0].get_attribute("href")
                            break
                    except Exception:
                        continue

            if target_url:
                driver.get(target_url)
            else:
                return

            # Refresh page if load time exceeds 5 seconds
            def refresh_page_if_slow():
                start_load_time = time.time()
                try:
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                except Exception:
                    # Refresh if page load takes more than 5 seconds
                    if time.time() - start_load_time > 5:
                        print(f"Page load took too long for product {product_code}, refreshing...")
                        driver.refresh()
                        time.sleep(2)  # Wait for the page to reload
                        driver.get(target_url)  # Retry loading the target URL
                        refresh_page_if_slow()  # Recursively call to ensure it reloads if needed
                return

            refresh_page_if_slow()  # Check and refresh if necessary

            # Download the largest image by size
            try:
                images = driver.find_elements(By.TAG_NAME, "img")
                largest_image = None
                largest_size = 0  # Used to track the largest image by area (width * height)

                for img in images:
                    img_src = img.get_attribute("src")
                    if not img_src:
                        # If the image doesn't have a 'src', check for the 'data-src' attribute
                        img_src = img.get_attribute("data-src")

                    if img_src:
                        try:
                            # Get the dimensions of the image on the page
                            width = int(img.get_attribute("width") or 0)
                            height = int(img.get_attribute("height") or 0)
                            
                            # Calculate the area of the image on the page (visible size, not file size)
                            image_area = width * height

                            # Check if this image is the largest so far
                            if image_area > largest_size:
                                largest_size = image_area
                                largest_image = img_src
                        except Exception:
                            continue  # Skip if there's an error

                if largest_image:
                    # Download the largest image
                    try:
                        image_data = requests.get(largest_image, headers=headers, allow_redirects=True, verify=False).content
                        image = Image.open(BytesIO(image_data))

                        # Save as jpg (for consistent filename extension)
                        img_extension = '.jpg' if not largest_image.lower().endswith('.webp') else '.webp'

                        # Save the image with the product code and extension
                        image.save(os.path.join(output_dir, f"{product_code}{img_extension}"))
                    except Exception as e:
                        print(f"Error downloading image from {largest_image}: {e}")
                else:
                    print(f"No image found for {product_code}.")
                    messagebox.showinfo(
                        "Image Download Error",
                        f"No valid image found for product code: {product_code}. Skipping this product."
                    )
            except Exception as e:
                print(f"Failed to download image for {product_code}: {e}")
                return

            processed_rows += 1
            elapsed_time = time.time() - start_time
            avg_time = elapsed_time / processed_rows
            eta = datetime.timedelta(seconds=int(avg_time * (total_rows - processed_rows)))

            # Update progress on the GUI in the main thread
            progress_label.config(text=f"Processed {processed_rows}/{total_rows} rows. ETA: {eta}")
            root.update_idletasks()

        # Process rows in a separate thread
        def process_rows():
            for row_num, row in enumerate(ws.iter_rows(
                min_row=start_row, max_row=end_row, min_col=ord(column)-64, max_col=ord(column)-64, values_only=True
            ), start=start_row):
                product_code = row[0]
                process_row(row_num, product_code)

            driver.quit()
            messagebox.showinfo("Completed", "Script finished successfully!")

        # Run processing in a separate thread to keep the GUI responsive
        threading.Thread(target=process_rows, daemon=True).start()

    # GUI setup
    root = Tk()
    root.title("Image Downloader")
    root.geometry("650x550")


    # Add the "Made by Artur Kuśmirek" text at the top
    Label(root, text="Made by Artur Kuśmirek, All rights reserved").pack()
    
    
    # File selection
    file_path = StringVar()
    Label(root, text="Excel File:").pack()
    Button(root, text="Browse", command=lambda: file_path.set(filedialog.askopenfilename())).pack()
    Label(root, textvariable=file_path).pack()

    # Column and row inputs
    column_choice = StringVar()
    Label(root, text="Column (A-Z):").pack()
    Entry(root, textvariable=column_choice).pack()

    start_row_var = StringVar(value="2")
    end_row_var = StringVar(value="10")
    Label(root, text="Start Row:").pack()
    Entry(root, textvariable=start_row_var).pack()
    Label(root, text="End Row:").pack()
    Entry(root, textvariable=end_row_var).pack()

    output_dir_var = StringVar()
    Label(root, text="Output Folder:").pack()
    Button(root, text="Browse", command=lambda: output_dir_var.set(filedialog.askdirectory())).pack()
    Label(root, textvariable=output_dir_var).pack()

    # Target website inputs
    target_website1_var = StringVar()
    target_website2_var = StringVar()
    target_website3_var = StringVar()
    target_website4_var = StringVar()

    Label(root, text="Target Website 1:").pack()
    Entry(root, textvariable=target_website1_var).pack()

    Label(root, text="Target Website 2:").pack()
    Entry(root, textvariable=target_website2_var).pack()

    Label(root, text="Target Website 3:").pack()
    Entry(root, textvariable=target_website3_var).pack()

    Label(root, text="Target Website 4:").pack()
    Entry(root, textvariable=target_website4_var).pack()

    # Start button
    Button(root, text="Start", command=start_script).pack()

    # Progress label
    progress_label = Label(root, text="Progress: 0/0")
    progress_label.pack()

    root.mainloop()

if __name__ == "__main__":
    main()